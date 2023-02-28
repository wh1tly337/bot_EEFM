import os
from datetime import datetime, timedelta

import openpyxl
from loguru import logger

from auxiliary.req_data import *


def all_cycle(filename, ender):
    """ Функция для улучшения читабельности кода и удобства """
    file_delete()
    file_renamer(filename, ender)


def file_delete():
    """ Функция удаляет старые файлы с расписанием"""
    try:
        os.remove(src_current_schedule)
    except Exception:
        pass


def file_renamer(filename, ender):
    """ Функция переименует получаемый файл"""
    try:
        os.rename(
            f"{src_files}{filename}.{ender}",
            f"{src_files}current_schedule.xlsx"
        )
    except Exception as ex:
        logger.error(ex)


async def get_schedule(doctor, time_period):
    """ Функция обработчик для получения расписания """
    try:
        wb_obj = openpyxl.load_workbook(src_current_schedule)  # обработка файла
        sheet_obj = wb_obj[doctor]  # выбор листа для работы (по фио)

        max_row = sheet_obj.max_row  # максимальное количество строк

        if time_period == 'today':  # получения расписания на сегодня
            text = await get_today_schedule(sheet_obj, max_row, 'today')
        elif time_period == 'week':  # получения расписания на неделю
            text = await get_weekly_schedule(sheet_obj, max_row)
        else:  # получения расписания на определенный день
            text = await get_today_schedule(sheet_obj, max_row, time_period)

        return text
    except Exception:
        logger.error(f"Доктора - {doctor} нет в расписании")
        text = f"Вас нет в листе с расписанием на эту неделю.\n" \
               f"Возможные способы решения данной проблемы:\n" \
               f"1) Вас не добавили в расписание\n" \
               f"2) При записи вашего ФИО произошла опечатка в листе с " \
               f"расписанием (обращаться к администратору), либо в базе " \
               f"данных бота (доступ у {director_name})"
        return text


async def get_today_schedule(sheet_obj, max_row, what_need):
    """ Функция для получения расписания на сегодня/определенный день недели """
    try:
        max_column = sheet_obj.max_column  # максимальное количество столбцов

        # для расписания на сегодня
        if what_need == 'today':
            day = datetime.now().strftime("%d-%m-%Y")
            text = f"Расписание на {datetime.now().strftime('%d-%m-%Y')}:\n\n"
        # для расписания на определенный день недели
        else:
            day = what_need
            text = f"Расписание на {what_need.lower()}:\n\n"

        cords = None  # переменная координат даты/дня недели в таблице
        counter = 0  # переменная считающая сколько дней нужно прибавить
        # цикл для получения координат крайней левой части таблицы каждого дня
        for i in range(1, max_column + 1):
            # список столбцов в которых находится ячейка с датой
            if what_need == 'today' and i in [1, 9, 17, 25, 33, 41, 49]:
                try:
                    # не брал просто значение ячейки потому что достается
                    # формула, а не итоговое значение
                    date = sheet_obj['A2'].value + timedelta(days=counter)
                    if date.strftime('%d-%m-%Y') == day:
                        cords = i
                    counter += 1
                except Exception as ex:
                    logger.error(ex)
            else:
                try:
                    # получаю значение дня недели
                    cell = sheet_obj.cell(row=2, column=i).value
                    if cell == day:
                        cords = i - 1
                except Exception:
                    continue

        result = []
        if cords is None:  # если нет расписания на текущую дату
            text += f"Нет расписания на этот день (необходимо сообщить " \
                    f"администратору об обновлении расписания)\n"
        else:
            # прохожусь по рядам самих таблиц
            for i in range(4, max_row + 1):
                none_counter = 0
                counter = 0
                pre_result = []
                # прохожусь по столбцам (от левой границы до правой)
                for j in range(cords, cords + 7):
                    cell = sheet_obj.cell(row=i, column=j).value
                    pre_result.append(cell)
                    if cell is None:
                        none_counter += 1
                    counter += 1
                    if counter == 7:
                        if none_counter != 7:
                            result.append(pre_result)
                        break
            length = len(result)
            if length == 0:
                text += f"Нет записей\n"
            else:
                for i in range(length):
                    text += f"{await text_formatter(i, result)}\n"

        return text
    except Exception as ex:
        logger.error(ex)


async def get_weekly_schedule(sheet_obj, max_row):
    """ Функция для получения расписания на неделю """
    try:
        counter, start, finish = 0, 1, 8
        result = []
        # цикл для прохода по каждому дню недели
        for _ in range(7):
            day_none_counter = 0
            # цикл для прохода по каждому ряду
            for i in range(2, max_row + 1):
                if i in [1, 2]:  # определяю конец таблицы каждого дня
                    finish = start + 2
                    f1 = 2
                else:
                    finish = start + 7
                    f1 = 7
                none_counter = 0
                pre_result = []
                # цикл для прохода по каждому столбцу
                for j in range(start, finish):
                    cell = sheet_obj.cell(row=i, column=j).value
                    pre_result.append(cell)
                    if cell is None:
                        none_counter += 1
                    counter += 1
                    if counter == f1:
                        if none_counter != f1:
                            result.append(pre_result)
                        else:
                            if none_counter == f1:
                                day_none_counter += 1
                            if day_none_counter == max_row - 3:
                                result.append('В этот день нет записей')
                        counter = 0
                        break
            start, finish = finish + 1, start + 7

        text = ''
        counter = 0
        # вариант одним сообщением
        # цикл для составления итогового текста сообщения с расписанием
        for i in range(len(result)):
            if len(result[i]) == 2:  # проверка строки для получения даты и дня
                date = sheet_obj['A2'].value + timedelta(days=counter)
                date = date.strftime('%d-%m-%Y')
                day = result[i][1]

                text += f"◉ {date} | {day}:\n\n"
                counter += 1
            else:
                if result[i][0] != 'Категория':  # проверка строки с заголовками
                    if result[i][0] != 'В':  # если в этот день нет записей
                        text += f"{await text_formatter(i, result)}\n"
                    else:  # если записи есть
                        text += f"{result[i]}\n\n"

        return text

        # вариант несколькими сообщениями
        # res = []
        #     # цикл для составления итогового текста сообщения с расписанием
        #     for i in range(len(result)):
        #         if len(result[i]) == 2:  # проверка на строку для получения
        #         даты и дня
        #             res.append(text)
        #             text = ''
        #
        #             date = sheet_obj['A2'].value + timedelta(days=counter)
        #             date = date.strftime('%d-%m-%Y')
        #             day = result[i][1]
        #
        #             text += f"◉ {date} | {day}:\n\n"
        #             counter += 1
        #         else:
        #             if result[i][0] != 'Категория':  # проверка на строку
        #             с заголовками
        #                 if result[i][0] != 'В':  # если в этот день нет
        #                 записей
        #                     text += f"{await text_formatter(i, result)}\n"
        #                 else:  # если записи есть
        #                     text += f"{result[i]}\n\n"
        #     res.append(text)
        #
        #     return res[1::]
    except Exception as ex:
        logger.error(ex)


async def text_formatter(i, result):
    """ Функция форматирования текста сообщения для расписания """
    inserts = [
        result[i][0],  # category
        result[i][1],  # time
        result[i][2],  # cabinet
        result[i][3],  # content
        result[i][4],  # comment
        result[i][5],  # card_number
        result[i][6],  # fio
    ]
    for i in range(7):
        if inserts[i] is None:
            inserts[i] = ' — '

    # noinspection PyUnresolvedReferences
    text = f"Категория пациента: {inserts[0]}\n" \
           f"Время записи: {inserts[1].strftime('%H:%M')}\n" \
           f"Номер кабинета: {inserts[2]}\n" \
           f"Соединение: {inserts[3]}\n" \
           f"Комментарий: {inserts[4]}\n" \
           f"Номер карты: {inserts[5]}\n" \
           f"ФИО пациента: {inserts[6]}\n"

    return text
